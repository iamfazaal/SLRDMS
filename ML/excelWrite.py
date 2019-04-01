# import openpyxl module 
import openpyxl 
from openpyxl.styles import Font
# Call a Workbook() function of openpyxl 
# to create a new blank Workbook object 
wb = openpyxl.Workbook() 

# Get workbook active sheet 
# from the active attribute 
sheet = wb.active 

# Cell objects also have row, column 
# and coordinate attributes that provide 
# location information for the cell. 

# Note: The first row or column integer 
# is 1, not 0. Cell object is created by 
# using sheet object's cell() method. 
c1 = sheet.cell(row = 1, column = 1)
sheet['A1'].font = Font(bold=True)
sheet['B1'].font = Font(bold=True)
sheet['C1'].font = Font(bold=True)

# writing values to cells 
c1.value = "Weight"

c2 = sheet.cell(row= 1 , column = 2) 
c2.value = "Texture"

c3 = sheet.cell(row= 1 , column = 3) 
c3.value = "Label"

# Once have a Worksheet object, one can 
# access a cell object by its name also. 
# A2 means column = 1 & row = 2. 
c4 = sheet['A2'] 
c4.value = "150g"

# B2 means column = 2 & row = 2. 
c5 = sheet['B2'] 
c5.value = "Bumpy"

c6 = sheet['C2'] 
c6.value = "Orange"

c7 = sheet['A3'] 
c7.value = "170g"

c8 = sheet['B3'] 
c8.value = "Bumpy"

c9 = sheet['C3'] 
c9.value = "Orange"

c10 = sheet['A4'] 
c10.value = "140g"

c11 = sheet['B4'] 
c11.value = "Smooth"

c12 = sheet['C4'] 
c12.value = "Apple"

c13 = sheet['A5'] 
c13.value = "130g"

c14 = sheet['B5'] 
c14.value = "Smooth"

c15 = sheet['C5'] 
c15.value = "Apple"

# Anytime you modify the Workbook object 
# or its sheets and cells, the spreadsheet 
# file will not be saved until you call 
# the save() workbook method. 
wb.save("C:\\Users\\nisal\\Documents\\Python\\demo.xlsx") 
